import subprocess
import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
import win32com.client as win32
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


script_path = r'C:\Users\user\Documents\codigos\Automacao_KM_Maxtrack\KM_CSV.py'
subprocess.run(['python', script_path], check=True)

server = 'server'
database = 'database'
username = 'user'
password = 'password'

conn_str = f'mssql+pyodbc://{username}:{password}@{server}/{database}?driver=SQL+Server'
engine = create_engine(conn_str)

file_path = r'C:\Users\user\Documents\SQL Server Management Studio\SELECT\VIEW TABELA KM ALERTA.sql'
with open(file_path, 'r') as file:
    query = file.read()

df = pd.read_sql(query, engine)

df['DATA_ENTRADA'] = pd.to_datetime(df['DATA_ENTRADA'], errors='coerce')
df['DATA_SAIDA'] = pd.to_datetime(df['DATA_SAIDA'], errors='coerce')
df['OPEGES'] = df['OPEGES'].fillna('').astype(str).str.upper()

df_sum = pd.read_excel(r'C:\Users\user\Downloads\resultado_placas.xlsx')
df_sum = df_sum.rename(columns={"Identificador/Placa": "PLACA", "Distância (Km)": "KM_CSV"})
df_merged = pd.merge(df, df_sum[['PLACA', 'KM_CSV']], on='PLACA', how='left')
df_merged['KM_CSV'] = df_merged['KM_CSV'].fillna('NULL')
df_merged['KM_DO_DIA_DA_ATUALIZAÇÃO'] = pd.to_numeric(df_merged['KM_CSV'], errors='coerce')

hoje = datetime.now().date()
hora_atual = datetime.now().time()

if hora_atual < datetime.strptime('15:00', '%H:%M').time():
    km_meta = 230
elif hora_atual > datetime.strptime('18:00', '%H:%M').time():
    km_meta = 450
else:
    km_meta = 230

grupos_excluidos = [
    "Cliente", 
    "Manutenção Interna", 
    "Manutenção Externa", 
    "Carregamento", 
    "Embarcador", 
    "Carregamento/Descarregamento", 
    "Descarregamento"
]

filtered_df = df_merged[
    (df_merged['KM_DO_DIA_DA_ATUALIZAÇÃO'] < km_meta) &
    ~(
        (df_merged['GRUPO'].isin(grupos_excluidos)) & 
        (df_merged['DATA_ENTRADA'].notna()) & 
        (
            (df_merged['DATA_SAIDA'].isna()) | 
            (
                (df_merged['DATA_SAIDA'].dt.date == hoje) & 
                (df_merged['DATA_SAIDA'].dt.time > pd.Timestamp('10:00').time()) & 
                (df_merged['OPEGES'] != 'GEISA.KARLA') 
            )
        )
    )
].copy()

filtered_df['JUSTIFICATIVA'] = ''
columns_to_drop = ['ULTIMA ATUALIZAÇÃO', 'ODOMETRO TOTAL', 'KM_DO_DIA_DA_ATUALIZAÇÃO']
filtered_df.drop(columns=columns_to_drop, inplace=True)
output_file = r'C:\Users\user\Documents\placas_nao_atendidas.xlsx'
filtered_df.to_excel(output_file, index=False)
wb = load_workbook(output_file)
ws = wb.active
ws.auto_filter.ref = ws.dimensions

lista_valores = [
    'Aguardando CTE', 'Aguardando Liberação de Carga/Descarga', 
    'Condições Adversas do Condutor', 'Condições Climáticas', 'DSR', 
    'Fiscalização ou Pesagem', 'Improdutividade do Condutor', 
    'Parada Não Programada (PNP)', 'Parada Programada (PP)', 
    'Problemas Mecânicos', 'Trânsito em Área Urbana', 
    'Trechos Curtos Entre Carregamento e Descarga', 
    'Aguardando Retorno do Programador', 
    'Viagem Realizada Dentro do Previsto', 'CT Disponível / Sem Cota', 
    'Manutenção', 'Petroreconcavo', 'Veículo em Adequação'
]

ws_lista = wb.create_sheet("Lista_Valores")
for i, valor in enumerate(lista_valores, start=1):
    ws_lista[f"A{i}"] = valor

intervalo_referencia = f"'Lista_Valores'!$A$1:$A${len(lista_valores)}"

validação_dados = DataValidation(
    type="list", 
    formula1=intervalo_referencia,
    allow_blank=True 
)

intervalo_células = "L2:L500"
validação_dados.add(intervalo_células)

ws.add_data_validation(validação_dados)

header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

ws_lista.sheet_state = 'veryHidden'

wb.save(output_file)

if not os.path.exists(output_file):
    print(f'O arquivo {output_file} não foi criado. Verifique o código.')
    exit()

outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)

email_body = f'''
Em anexo está a tabela com as placas que não atenderam às metas.
A meta de quilometragem para este horário é de {km_meta} km.

Grupos excluídos da filtragem se tiverem DATA_ENTRADA e não tiverem DATA_SAIDA:
{', '.join(grupos_excluidos)}

Requisitos adicionais:
- Placas com DATA_SAIDA igual à data atual e horário após 10:00 não serão filtradas.
'''
mail.Subject = 'Relatório de Placas Não Atendidas'
mail.Body = email_body
mail.To = 'email'
mail.Attachments.Add(output_file)

mail.Send()
print('E-mail enviado com sucesso!')
